import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

# Read the JSON file
with open('Measures/measures.json', 'r') as f:
    data = json.load(f)

# Organize data by lab and workstation
organized_data = defaultdict(lambda: defaultdict(list))

for entry in data['data']:
    position = entry['position']
    # Extract lab and workstation from position (e.g., "LAP1-0" -> lab="LAP1", workstation="0")
    parts = position.split('-')
    lab = parts[0]
    workstation = parts[1] if len(parts) > 1 else "Unknown"
    
    organized_data[lab][workstation].append(entry)

# Create Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Style definitions
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
lab_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
lab_font = Font(bold=True)

# Create a sheet for each lab
for lab in sorted(organized_data.keys()):
    ws = wb.create_sheet(title=lab)
    
    row = 1
    
    # Add lab title
    ws.merge_cells(f'A{row}:F{row}')
    title_cell = ws[f'A{row}']
    title_cell.value = f"Lab: {lab}"
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    # Add headers for each workstation
    for workstation in sorted(organized_data[lab].keys()):
        # Workstation header
        ws.merge_cells(f'A{row}:F{row}')
        ws_cell = ws[f'A{row}']
        ws_cell.value = f"Workstation: {workstation}"
        ws_cell.fill = lab_fill
        ws_cell.font = lab_font
        ws_cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        
        # Column headers
        headers = ["ID", "Position", "Temperature (°C)", "Humidity (%)", "Luminosity", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Add data rows for this workstation
        for entry in organized_data[lab][workstation]:
            ws.cell(row=row, column=1).value = entry['id']
            ws.cell(row=row, column=2).value = entry['position']
            ws.cell(row=row, column=3).value = entry['temperature']
            ws.cell(row=row, column=4).value = entry['humidity']
            ws.cell(row=row, column=5).value = entry['luminosity']
            ws.cell(row=row, column=6).value = entry['timestamp']
            row += 1
        
        row += 1  # Add blank row between workstations

# Create a summary sheet with all measures in columns
summary_ws = wb.create_sheet(title="All Measures", index=0)

# Column headers for summary sheet
headers = ["ID", "Position", "Temperature (°C)", "Humidity (%)", "Luminosity", "Timestamp"]
for col, header in enumerate(headers, 1):
    cell = summary_ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add all data rows
row = 2
for entry in data['data']:
    summary_ws.cell(row=row, column=1).value = entry['id']
    summary_ws.cell(row=row, column=2).value = entry['position']
    summary_ws.cell(row=row, column=3).value = entry['temperature']
    summary_ws.cell(row=row, column=4).value = entry['humidity']
    summary_ws.cell(row=row, column=5).value = entry['luminosity']
    summary_ws.cell(row=row, column=6).value = entry['timestamp']
    row += 1

# Adjust column widths for all sheets
for ws in wb.sheetnames:
    worksheet = wb[ws]
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['C'].width = 16
    worksheet.column_dimensions['D'].width = 14
    worksheet.column_dimensions['E'].width = 14
    worksheet.column_dimensions['F'].width = 20

# Save the workbook
output_path = 'Measures/weather_data.xlsx'
wb.save(output_path)
print(f"Excel file created successfully: {output_path}")
